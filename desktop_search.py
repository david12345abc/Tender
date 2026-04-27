"""Десктопное приложение для поиска актуальных процедур на ЭТП ГПБ.

Работает через уже авторизованный Chrome с DevTools на порту 9222.
Chrome стартует только по клику «Поиск» (`start_chrome.ps1`), а до этого
приложение не трогает браузер. Все сетевые операции идут в отдельном
QThread, так что UI не блокируется и не падает при ошибках.

Запуск:
    python desktop_search.py

Требует: Python 3.10+, PySide6, selenium, openpyxl (для экспорта), и
наличие `start_chrome.ps1` рядом со скриптом.
"""
from __future__ import annotations

import json
import sys
import traceback
import webbrowser
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Optional

from PySide6.QtCore import (
    QAbstractTableModel,
    QDate,
    QModelIndex,
    QObject,
    QSortFilterProxyModel,
    Qt,
    QThread,
    QTimer,
    Signal,
    Slot,
)
from PySide6.QtGui import QAction, QColor, QFont, QKeySequence, QPalette
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QComboBox,
    QDateEdit,
    QDoubleSpinBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMenu,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QScrollArea,
    QSpinBox,
    QSplitter,
    QStatusBar,
    QTableView,
    QVBoxLayout,
    QWidget,
)

from etp_client import (
    HARD_SERVER_LIMIT,
    EtpClient,
    STEP_ID_LABELS,
    TREND_PUR_LABELS,
    step_id_label,
    trend_pur_label,
)

APP_TITLE = "ЭТП ГПБ — поиск тендеров"
CACHE_FILE = Path(__file__).parent / "cache" / "desktop_search_cache.json"
VIEW_URL = "https://etpgaz.gazprombank.ru/#com/procedure/view/id/{pid}"

COLUMNS: list[tuple[str, str]] = [
    ("registry_number", "Реестровый №"),
    ("trend_pur_label", "Тип"),
    ("organizer", "Организатор"),
    ("title", "Наименование"),
    ("tags_label", "Теги"),
    ("applics_count", "Намерений"),
    ("date_end_registration", "Приём заявок до"),
    ("total_price", "Сумма"),
    ("step_label", "Статус"),
]


# -----------------------------------------------------------------------------
# Утилиты парсинга / форматирования
# -----------------------------------------------------------------------------

def _parse_dt(value: Any) -> Optional[datetime]:
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    if not s:
        return None
    if len(s) >= 25 and (s[-6] in "+-") and s[-3] == ":":
        s = s[:-3] + s[-2:]
    for fmt in (
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%d %H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s)
    except Exception:
        return None


def _parse_price(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", ".").replace(" ", ""))
    except (TypeError, ValueError):
        return None


def _fmt_money(v: Optional[float], currency: str = "RUB") -> str:
    if v is None:
        return ""
    return f"{v:,.2f}".replace(",", " ") + (f" {currency}" if currency else "")


def _fmt_date(dt: Optional[datetime]) -> str:
    if dt is None:
        return ""
    return (
        dt.strftime("%d.%m.%Y %H:%M")
        if (dt.hour or dt.minute)
        else dt.strftime("%d.%m.%Y")
    )


# -----------------------------------------------------------------------------
# Параметры
# -----------------------------------------------------------------------------

@dataclass
class SearchParams:
    date_from: str = ""
    query: str = ""
    tag_id: Optional[int] = None
    sort: str = "id"
    direction: str = "DESC"


@dataclass
class ClientFilters:
    registry_contains: str = ""
    organizer_contains: str = ""
    title_contains: str = ""
    trend_pur: str = ""
    step_id: str = ""
    applics_min: Optional[int] = None
    applics_max: Optional[int] = None
    price_min: Optional[float] = None
    price_max: Optional[float] = None
    end_from: Optional[datetime] = None
    end_to: Optional[datetime] = None


# -----------------------------------------------------------------------------
# Модели данных
# -----------------------------------------------------------------------------

class ProcedureTableModel(QAbstractTableModel):
    COL_KEYS = [c[0] for c in COLUMNS]
    COL_TITLES = [c[1] for c in COLUMNS]

    def __init__(self, parent: Optional[QObject] = None) -> None:
        super().__init__(parent)
        self._rows: list[dict[str, Any]] = []

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:
        return 0 if parent.isValid() else len(self._rows)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:
        return 0 if parent.isValid() else len(self.COL_KEYS)

    def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.DisplayRole) -> Any:
        if role == Qt.DisplayRole and orientation == Qt.Horizontal:
            return self.COL_TITLES[section]
        if role == Qt.DisplayRole and orientation == Qt.Vertical:
            return section + 1
        return None

    def _display(self, proc: dict[str, Any], key: str) -> Any:
        if key == "trend_pur_label":
            return trend_pur_label(proc.get("trend_pur"))
        if key == "step_label":
            return step_id_label(proc.get("step_id"))
        if key == "organizer":
            return proc.get("short_name") or proc.get("full_name") or ""
        if key == "tags_label":
            tags = proc.get("tags") or []
            return ", ".join(str(t) for t in tags) if tags else ""
        if key == "date_end_registration":
            return _fmt_date(_parse_dt(proc.get("date_end_registration")))
        if key == "total_price":
            p = _parse_price(proc.get("total_price"))
            return _fmt_money(p, proc.get("currency_name") or "RUB")
        if key == "applics_count":
            return proc.get("applics_count") if proc.get("applics_count") is not None else ""
        if key == "title":
            return proc.get("title") or ""
        if key == "registry_number":
            return proc.get("registry_number") or proc.get("procedure_number") or ""
        return proc.get(key, "")

    def _sort_value(self, proc: dict[str, Any], key: str) -> Any:
        if key == "total_price":
            return _parse_price(proc.get("total_price")) or 0.0
        if key == "applics_count":
            return int(proc.get("applics_count") or 0)
        if key == "date_end_registration":
            return _parse_dt(proc.get("date_end_registration")) or datetime.min
        if key == "trend_pur_label":
            return str(proc.get("trend_pur") or "")
        if key == "step_label":
            return str(proc.get("step_id") or "")
        if key == "organizer":
            return str(proc.get("short_name") or proc.get("full_name") or "").lower()
        if key == "title":
            return str(proc.get("title") or "").lower()
        if key == "registry_number":
            return str(proc.get("registry_number") or proc.get("procedure_number") or "")
        return str(proc.get(key) or "")

    def data(self, index: QModelIndex, role: int = Qt.DisplayRole) -> Any:
        if not index.isValid():
            return None
        row = index.row()
        if not (0 <= row < len(self._rows)):
            return None
        proc = self._rows[row]
        col_key = self.COL_KEYS[index.column()]

        if role == Qt.DisplayRole:
            return self._display(proc, col_key)
        if role == Qt.UserRole:
            return self._sort_value(proc, col_key)
        if role == Qt.TextAlignmentRole:
            if col_key in ("total_price", "applics_count"):
                return int(Qt.AlignRight | Qt.AlignVCenter)
            return int(Qt.AlignLeft | Qt.AlignVCenter)
        if role == Qt.ToolTipRole:
            if col_key == "organizer":
                parts: list[str] = []
                if proc.get("full_name"):
                    parts.append(str(proc["full_name"]))
                if proc.get("org_inn"):
                    parts.append(f"ИНН {proc['org_inn']}")
                if proc.get("org_kpp"):
                    parts.append(f"КПП {proc['org_kpp']}")
                return "\n".join(parts) or None
            if col_key == "title":
                return str(proc.get("title") or "")
            if col_key == "date_end_registration":
                return str(proc.get("date_end_registration") or "")
            if col_key == "registry_number":
                info = []
                for k in ("registry_number", "procedure_number", "procedure_number2"):
                    v = proc.get(k)
                    if v:
                        info.append(f"{k}: {v}")
                info.append(f"id: {proc.get('id')}")
                return "\n".join(info)
        if role == Qt.BackgroundRole:
            step = proc.get("step_id")
            if step == "registration":
                return QColor(225, 245, 225)
            if step == "applic_access":
                return QColor(230, 240, 255)
            if step == "second_parts":
                return QColor(250, 240, 210)
            if step == "finalizing_procedure":
                return QColor(235, 235, 235)
        return None

    def set_rows(self, procs: list[dict[str, Any]]) -> None:
        self.beginResetModel()
        self._rows = list(procs)
        self.endResetModel()

    def append_rows(self, procs: list[dict[str, Any]]) -> None:
        if not procs:
            return
        first = len(self._rows)
        self.beginInsertRows(QModelIndex(), first, first + len(procs) - 1)
        self._rows.extend(procs)
        self.endInsertRows()

    def clear(self) -> None:
        self.beginResetModel()
        self._rows = []
        self.endResetModel()

    def rows(self) -> list[dict[str, Any]]:
        return self._rows

    def row_at(self, row: int) -> Optional[dict[str, Any]]:
        if 0 <= row < len(self._rows):
            return self._rows[row]
        return None


class ProcedureFilterProxy(QSortFilterProxyModel):
    def __init__(self, parent: Optional[QObject] = None) -> None:
        super().__init__(parent)
        self._flt = ClientFilters()
        self.setSortRole(Qt.UserRole)
        self.setFilterCaseSensitivity(Qt.CaseInsensitive)

    def set_filters(self, flt: ClientFilters) -> None:
        self._flt = flt
        self.invalidateFilter()

    def filterAcceptsRow(self, source_row: int, source_parent: QModelIndex) -> bool:
        model = self.sourceModel()
        if not isinstance(model, ProcedureTableModel):
            return True
        proc = model.row_at(source_row)
        if proc is None:
            return False
        f = self._flt

        if f.registry_contains:
            needle = f.registry_contains.lower()
            blob = " ".join(
                str(proc.get(k) or "").lower()
                for k in ("registry_number", "procedure_number", "procedure_number2")
            )
            if needle not in blob:
                return False

        if f.organizer_contains:
            needle = f.organizer_contains.lower()
            blob = " ".join(
                str(proc.get(k) or "").lower()
                for k in ("short_name", "full_name", "org_inn", "org_ogrn")
            )
            if needle not in blob:
                return False

        if f.title_contains:
            if f.title_contains.lower() not in str(proc.get("title") or "").lower():
                return False

        if f.trend_pur and str(proc.get("trend_pur") or "") != f.trend_pur:
            return False
        if f.step_id and str(proc.get("step_id") or "") != f.step_id:
            return False

        apc = int(proc.get("applics_count") or 0)
        if f.applics_min is not None and apc < f.applics_min:
            return False
        if f.applics_max is not None and apc > f.applics_max:
            return False

        price = _parse_price(proc.get("total_price"))
        if f.price_min is not None and (price is None or price < f.price_min):
            return False
        if f.price_max is not None and (price is None or price > f.price_max):
            return False

        if f.end_from is not None or f.end_to is not None:
            end_dt = _parse_dt(proc.get("date_end_registration"))
            if end_dt is None:
                return False
            naive = end_dt.replace(tzinfo=None)
            if f.end_from is not None and naive < f.end_from:
                return False
            if f.end_to is not None and naive > f.end_to:
                return False
        return True


# -----------------------------------------------------------------------------
# Worker в отдельном потоке
# -----------------------------------------------------------------------------

class Worker(QObject):
    """Универсальный работник: выполняет одну задачу за жизнь.

    Сигналы:
        progress(str)      — сообщения о прогрессе
        session(bool, str) — результат проверки сессии
        batch(list, int, int) — загружена пачка: procedures, start, total
        error(str)         — неперехваченное исключение
        finished()         — всегда вызывается после run()
    """

    progress = Signal(str)
    session = Signal(bool, str)
    batch = Signal(list, int, int)
    error = Signal(str)
    finished = Signal()

    def __init__(self, fn: Callable[["Worker"], None]) -> None:
        super().__init__()
        self._fn = fn
        self._stop = False

    def request_stop(self) -> None:
        self._stop = True

    def is_stop_requested(self) -> bool:
        return self._stop

    @Slot()
    def run(self) -> None:
        try:
            self._fn(self)
        except Exception as e:
            tb = traceback.format_exc()
            self.error.emit(f"{type(e).__name__}: {e}\n{tb}")
        finally:
            self.finished.emit()


class TaskRunner(QObject):
    """Запускает `Worker` в отдельном QThread. Гарантирует корректное завершение."""

    def __init__(self, parent: Optional[QObject] = None) -> None:
        super().__init__(parent)
        self._thread: Optional[QThread] = None
        self._worker: Optional[Worker] = None

    def is_running(self) -> bool:
        return self._thread is not None

    def start(
        self,
        fn: Callable[[Worker], None],
        on_progress: Optional[Callable[[str], None]] = None,
        on_session: Optional[Callable[[bool, str], None]] = None,
        on_batch: Optional[Callable[[list, int, int], None]] = None,
        on_error: Optional[Callable[[str], None]] = None,
        on_done: Optional[Callable[[], None]] = None,
    ) -> Worker:
        if self._thread is not None:
            raise RuntimeError("Task already running")

        thread = QThread(self.parent())
        worker = Worker(fn)
        worker.moveToThread(thread)

        if on_progress:
            worker.progress.connect(on_progress)
        if on_session:
            worker.session.connect(on_session)
        if on_batch:
            worker.batch.connect(on_batch)
        if on_error:
            worker.error.connect(on_error)

        worker.finished.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)

        def _cleanup() -> None:
            self._thread = None
            self._worker = None
            if on_done:
                try:
                    on_done()
                except Exception:
                    traceback.print_exc()

        thread.finished.connect(_cleanup)
        thread.started.connect(worker.run)

        self._thread = thread
        self._worker = worker
        thread.start()
        return worker

    def request_stop(self) -> None:
        if self._worker:
            self._worker.request_stop()

    def shutdown(self, wait_ms: int = 3000) -> None:
        if self._worker:
            self._worker.request_stop()
        if self._thread:
            self._thread.quit()
            self._thread.wait(wait_ms)
        self._thread = None
        self._worker = None


# -----------------------------------------------------------------------------
# Задачи (запускаются внутри Worker)
# -----------------------------------------------------------------------------

def _make_search_task(
    client: EtpClient,
    params: SearchParams,
    start: int,
    batches_left: int,
) -> Callable[[Worker], None]:
    """Задача: запустить Chrome (если надо), подключиться, проверить сессию,
    скачать одну или несколько пачек.

    batches_left — сколько пачек подряд скачать. 1 = «одна». 9999 = «всё».
    """

    def _run(w: Worker) -> None:
        if w.is_stop_requested():
            return

        if not client.is_chrome_running():
            w.progress.emit("Запускаю Chrome с DevTools…")
            try:
                client.ensure_chrome(timeout=45)
            except Exception as e:
                w.error.emit(f"Не удалось запустить Chrome: {e}")
                return
        if w.is_stop_requested():
            return

        if client.driver is None:
            w.progress.emit("Подключаюсь к Chrome DevTools…")
            try:
                client.connect()
            except Exception as e:
                w.error.emit(f"Ошибка подключения к Chrome: {e}")
                return

        if w.is_stop_requested():
            return

        w.progress.emit("Получаю CSRF-токен…")
        try:
            client.pull_token()
        except Exception:
            pass

        if w.is_stop_requested():
            return

        cur_start = start
        loaded_this_task = 0
        total: Optional[int] = None
        batches_done = 0

        while batches_done < batches_left:
            if w.is_stop_requested():
                return
            w.progress.emit(
                f"Запрос Procedure.list: start={cur_start}, limit={HARD_SERVER_LIMIT}"
                + (f"  (загружено {loaded_this_task}/{total})" if total else "")
            )
            res = client.fetch_page(
                start=cur_start,
                limit=HARD_SERVER_LIMIT,
                date_from=params.date_from or None,
                query=params.query or None,
                tag_id=params.tag_id,
                sort=params.sort,
                direction=params.direction,
            )
            if res.get("error"):
                err_text = str(res["error"])
                err_low = err_text.lower()
                if (
                    "no such window" in err_low
                    or "web view not found" in err_low
                    or "target window already closed" in err_low
                    or "target frame detached" in err_low
                    or "invalid session id" in err_low
                ):
                    short = (
                        "Вкладка ЭТП была закрыта в Chrome. "
                        "Открыл её заново — попробуйте ещё раз нажать «Поиск»."
                    )
                    w.error.emit(short)
                else:
                    w.error.emit(f"Сервер вернул ошибку: {err_text}")
                return
            if res.get("no_access") or res.get("no_session"):
                msg = res.get("message") or "Нет доступа / сессия не активна."
                w.session.emit(
                    False,
                    f"{msg}\n\nВ Chrome: «Войти» → «ЕСИА + ЭП» → пройдите до конца, "
                    "затем снова нажмите «Поиск».",
                )
                return
            procs = res.get("procedures") or []
            if total is None:
                total = int(res.get("totalCount") or 0)
            w.batch.emit(procs, cur_start, total or 0)
            loaded_this_task += len(procs)
            batches_done += 1
            if not procs:
                break
            if total and cur_start + len(procs) >= total:
                break
            cur_start += len(procs)

        w.session.emit(True, "Готово.")

    return _run


# -----------------------------------------------------------------------------
# UI-виджеты
# -----------------------------------------------------------------------------

APP_STYLE = """
QMainWindow { background-color: #f7f8fa; }

#TopBar {
    background-color: #ffffff;
    border-bottom: 1px solid #e3e5e8;
}
#TopBarTitle {
    font-size: 16px;
    font-weight: 700;
    color: #1a1d22;
}
#TopBarSubtitle {
    font-size: 11px;
    color: #6a717a;
}
#SessionBadge {
    padding: 4px 10px;
    border-radius: 10px;
    font-weight: 600;
}
#SessionBadge[ok="true"]  { background-color: #e6f6ea; color: #1f7a3a; }
#SessionBadge[ok="false"] { background-color: #fbe9e9; color: #a61b1b; }
#SessionBadge[ok="idle"]  { background-color: #eef0f4; color: #4a515a; }

#Sidebar {
    background-color: #ffffff;
    border-right: 1px solid #e3e5e8;
}
#SidebarBody {
    background-color: #ffffff;
}
#SidebarTitle {
    font-size: 13px;
    font-weight: 700;
    color: #1a1d22;
    padding: 2px 0;
}
#SidebarSection {
    color: #6a717a;
    font-size: 11px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.3px;
    padding-top: 8px;
}

QLabel {
    color: #1a1d22;
}

QLineEdit, QComboBox, QDateEdit, QSpinBox, QDoubleSpinBox {
    padding: 5px 6px;
    border: 1px solid #d4d8dd;
    border-radius: 6px;
    background-color: #ffffff;
    color: #1a1d22;
    selection-background-color: #cfe0ff;
    selection-color: #1a1d22;
    min-height: 30px;
}
QLineEdit:focus, QComboBox:focus, QDateEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus {
    border-color: #3572e0;
}
QLineEdit:disabled, QComboBox:disabled, QDateEdit:disabled, QSpinBox:disabled, QDoubleSpinBox:disabled {
    color: #7b818a;
    background-color: #f3f5f7;
}
QAbstractSpinBox::up-button, QAbstractSpinBox::down-button {
    width: 18px;
    background-color: #f7f8fa;
}
QAbstractSpinBox::up-arrow, QAbstractSpinBox::down-arrow {
    width: 8px;
    height: 8px;
}

QPushButton {
    padding: 7px 14px;
    border: 1px solid #d4d8dd;
    border-radius: 6px;
    background-color: #ffffff;
    color: #1a1d22;
}
QPushButton:hover { background-color: #f0f3f7; }
QPushButton:disabled { color: #a2a8b0; background-color: #f5f6f8; }

QPushButton#Primary {
    background-color: #3572e0;
    border: 1px solid #2a5bc0;
    color: white;
    font-weight: 600;
}
QPushButton#Primary:hover     { background-color: #2a5bc0; }
QPushButton#Primary:disabled  { background-color: #9fbaed; border-color: #9fbaed; }

QPushButton#Danger {
    background-color: #ffffff;
    border: 1px solid #e0b4b4;
    color: #a61b1b;
}
QPushButton#Danger:hover { background-color: #fbe9e9; }

QTableView {
    background-color: #ffffff;
    alternate-background-color: #f8f9fb;
    gridline-color: #eceff3;
    selection-background-color: #cfe0ff;
    selection-color: #1a1d22;
    border: 1px solid #e3e5e8;
}
QHeaderView::section {
    background-color: #f0f3f7;
    padding: 6px 8px;
    border: 0px;
    border-right: 1px solid #e3e5e8;
    border-bottom: 1px solid #e3e5e8;
    font-weight: 600;
    color: #3a4048;
}
QTableView::item { padding: 4px 8px; }

QStatusBar { background: #ffffff; border-top: 1px solid #e3e5e8; }
QProgressBar {
    border: 1px solid #d4d8dd;
    border-radius: 5px;
    background-color: #ffffff;
    text-align: center;
    height: 14px;
}
QProgressBar::chunk { background-color: #3572e0; border-radius: 4px; }

QCheckBox { spacing: 5px; }

QToolTip {
    background-color: #2c2f36;
    color: #ffffff;
    border: 1px solid #2c2f36;
    padding: 4px 6px;
    border-radius: 4px;
}

#FilterLabel {
    color: #3a4048;
    font-size: 11px;
    font-weight: 600;
    margin-top: 4px;
}

#BottomBar {
    background-color: #ffffff;
    border: 1px solid #e3e5e8;
    border-radius: 8px;
}
"""


class Sidebar(QWidget):
    """Левая панель со всеми фильтрами и основными действиями."""

    searchRequested = Signal()
    resetRequested = Signal()
    clientFiltersChanged = Signal()
    loadMoreRequested = Signal()
    loadAllRequested = Signal()
    stopRequested = Signal()

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setObjectName("Sidebar")
        self.setMinimumWidth(380)
        self.setMaximumWidth(460)
        self._build_ui()

    def _build_ui(self) -> None:
        scroll = QScrollArea(self)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        body = QWidget()
        body.setObjectName("SidebarBody")
        body_layout = QVBoxLayout(body)
        body_layout.setContentsMargins(14, 12, 14, 12)
        body_layout.setSpacing(8)

        title = QLabel("Фильтры поиска")
        title.setObjectName("SidebarTitle")
        body_layout.addWidget(title)

        # ===== Серверные фильтры =====
        body_layout.addWidget(self._section("Серверный поиск"))

        date_lbl = QLabel("Дата публикации с")
        date_lbl.setObjectName("FilterLabel")
        self.ed_date_from = QDateEdit()
        self.ed_date_from.setCalendarPopup(True)
        self.ed_date_from.setDisplayFormat("dd.MM.yyyy")
        self.ed_date_from.setMinimumWidth(180)
        self.ed_date_from.setDate(QDate.currentDate().addYears(-1))
        body_layout.addWidget(date_lbl)
        body_layout.addWidget(self.ed_date_from)

        qlbl = QLabel("Поиск по наименованию (целиком/фраза)")
        qlbl.setObjectName("FilterLabel")
        qlbl.setWordWrap(True)
        self.ed_query = QLineEdit()
        self.ed_query.setPlaceholderText("например: поставка трубы")
        body_layout.addWidget(qlbl)
        body_layout.addWidget(self.ed_query)

        tag_lbl = QLabel("Тег (id, 0 = любой)")
        tag_lbl.setObjectName("FilterLabel")
        self.ed_tag_id = QSpinBox()
        self.ed_tag_id.setRange(0, 1_000_000)
        self.ed_tag_id.setSpecialValueText("— любой —")
        self.ed_tag_id.setMinimumWidth(180)
        body_layout.addWidget(tag_lbl)
        body_layout.addWidget(self.ed_tag_id)

        # Основная кнопка поиска
        self.btn_search = QPushButton("Поиск")
        self.btn_search.setObjectName("Primary")
        self.btn_search.setMinimumHeight(34)
        body_layout.addSpacing(4)
        body_layout.addWidget(self.btn_search)

        # ===== Клиентские фильтры =====
        body_layout.addWidget(self._section("Фильтры результатов"))
        hint = QLabel("применяются мгновенно к уже загруженному списку")
        hint.setStyleSheet("color: #6a717a; font-size: 10px;")
        hint.setWordWrap(True)
        body_layout.addWidget(hint)

        filters_wrap = QVBoxLayout()
        filters_wrap.setContentsMargins(0, 0, 0, 0)
        filters_wrap.setSpacing(6)

        reg_lbl = QLabel("Реестровый №")
        reg_lbl.setObjectName("FilterLabel")
        self.ed_registry = QLineEdit()
        self.ed_registry.setPlaceholderText("содержит …")
        filters_wrap.addWidget(reg_lbl)
        filters_wrap.addWidget(self.ed_registry)

        org_lbl = QLabel("Организатор")
        org_lbl.setObjectName("FilterLabel")
        self.ed_organizer = QLineEdit()
        self.ed_organizer.setPlaceholderText("название / ИНН / ОГРН")
        filters_wrap.addWidget(org_lbl)
        filters_wrap.addWidget(self.ed_organizer)

        title_lbl = QLabel("Наименование")
        title_lbl.setObjectName("FilterLabel")
        self.ed_title_local = QLineEdit()
        self.ed_title_local.setPlaceholderText("слова в названии")
        filters_wrap.addWidget(title_lbl)
        filters_wrap.addWidget(self.ed_title_local)

        trend_lbl = QLabel("Тип")
        trend_lbl.setObjectName("FilterLabel")
        self.cb_trend = QComboBox()
        self.cb_trend.addItem("— все —", "")
        for code, lbl in TREND_PUR_LABELS.items():
            self.cb_trend.addItem(lbl, code)
        filters_wrap.addWidget(trend_lbl)
        filters_wrap.addWidget(self.cb_trend)

        step_lbl = QLabel("Статус")
        step_lbl.setObjectName("FilterLabel")
        self.cb_step = QComboBox()
        self.cb_step.addItem("— все —", "")
        # Удалим дубли лейблов «Рассмотрение вторых частей»
        added_labels: set[str] = set()
        for code, lbl in STEP_ID_LABELS.items():
            if lbl in added_labels:
                continue
            added_labels.add(lbl)
            self.cb_step.addItem(lbl, code)
        filters_wrap.addWidget(step_lbl)
        filters_wrap.addWidget(self.cb_step)

        # Намерений
        apc_lbl = QLabel("Намерений")
        apc_lbl.setObjectName("FilterLabel")
        apc_row = QWidget()
        apc_lay = QHBoxLayout(apc_row)
        apc_lay.setContentsMargins(0, 0, 0, 0)
        apc_lay.setSpacing(4)
        self.sb_apc_min = QSpinBox()
        self.sb_apc_min.setRange(0, 1_000_000)
        self.sb_apc_min.setSpecialValueText("от")
        self.sb_apc_min.setMinimumWidth(120)
        self.sb_apc_max = QSpinBox()
        self.sb_apc_max.setRange(0, 1_000_000)
        self.sb_apc_max.setSpecialValueText("до")
        self.sb_apc_max.setMinimumWidth(120)
        apc_lay.addWidget(self.sb_apc_min)
        apc_lay.addWidget(QLabel("—"))
        apc_lay.addWidget(self.sb_apc_max)
        filters_wrap.addWidget(apc_lbl)
        filters_wrap.addWidget(apc_row)

        # Сумма
        price_lbl = QLabel("Сумма, RUB")
        price_lbl.setObjectName("FilterLabel")
        price_row = QWidget()
        price_lay = QHBoxLayout(price_row)
        price_lay.setContentsMargins(0, 0, 0, 0)
        price_lay.setSpacing(4)
        self.sb_price_min = QDoubleSpinBox()
        self.sb_price_min.setDecimals(2)
        self.sb_price_min.setRange(0, 1e13)
        self.sb_price_min.setGroupSeparatorShown(True)
        self.sb_price_min.setSpecialValueText("от")
        self.sb_price_min.setMinimumWidth(140)
        self.sb_price_max = QDoubleSpinBox()
        self.sb_price_max.setDecimals(2)
        self.sb_price_max.setRange(0, 1e13)
        self.sb_price_max.setGroupSeparatorShown(True)
        self.sb_price_max.setSpecialValueText("до")
        self.sb_price_max.setMinimumWidth(140)
        price_lay.addWidget(self.sb_price_min)
        price_lay.addWidget(QLabel("—"))
        price_lay.addWidget(self.sb_price_max)
        filters_wrap.addWidget(price_lbl)
        filters_wrap.addWidget(price_row)

        # Приём заявок
        end_lbl = QLabel("Приём заявок")
        end_lbl.setObjectName("FilterLabel")
        self.cb_end_from_en = QCheckBox("с")
        self.de_end_from = QDateEdit()
        self.de_end_from.setCalendarPopup(True)
        self.de_end_from.setDisplayFormat("dd.MM.yyyy")
        self.de_end_from.setMinimumWidth(150)
        self.de_end_from.setDate(QDate.currentDate())
        self.de_end_from.setEnabled(False)
        self.cb_end_from_en.toggled.connect(self.de_end_from.setEnabled)

        self.cb_end_to_en = QCheckBox("по")
        self.de_end_to = QDateEdit()
        self.de_end_to.setCalendarPopup(True)
        self.de_end_to.setDisplayFormat("dd.MM.yyyy")
        self.de_end_to.setMinimumWidth(150)
        self.de_end_to.setDate(QDate.currentDate().addMonths(3))
        self.de_end_to.setEnabled(False)
        self.cb_end_to_en.toggled.connect(self.de_end_to.setEnabled)

        end_grid = QGridLayout()
        end_grid.setContentsMargins(0, 0, 0, 0)
        end_grid.setHorizontalSpacing(6)
        end_grid.setVerticalSpacing(6)
        end_grid.addWidget(self.cb_end_from_en, 0, 0)
        end_grid.addWidget(self.de_end_from, 0, 1)
        end_grid.addWidget(self.cb_end_to_en, 1, 0)
        end_grid.addWidget(self.de_end_to, 1, 1)
        filters_wrap.addWidget(end_lbl)
        filters_wrap.addLayout(end_grid)

        body_layout.addLayout(filters_wrap)

        # Кнопка сброса фильтров
        self.btn_reset = QPushButton("Сбросить фильтры")
        body_layout.addWidget(self.btn_reset)

        body_layout.addStretch(1)
        scroll.setWidget(body)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll)

        # Подключение сигналов клиентских фильтров (мгновенные)
        for w in (self.ed_registry, self.ed_organizer, self.ed_title_local):
            w.returnPressed.connect(self.clientFiltersChanged)
            w.editingFinished.connect(self.clientFiltersChanged)
        for w in (self.cb_trend, self.cb_step):
            w.currentIndexChanged.connect(lambda *_: self.clientFiltersChanged.emit())
        for w in (self.sb_apc_min, self.sb_apc_max, self.sb_price_min, self.sb_price_max):
            w.editingFinished.connect(self.clientFiltersChanged)
        for w in (self.cb_end_from_en, self.cb_end_to_en):
            w.toggled.connect(lambda *_: self.clientFiltersChanged.emit())
        for w in (self.de_end_from, self.de_end_to):
            w.dateChanged.connect(lambda *_: self.clientFiltersChanged.emit())

        self.btn_search.clicked.connect(self.searchRequested)
        self.btn_reset.clicked.connect(self.resetRequested)

    def _section(self, text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setObjectName("SidebarSection")
        return lbl

    def search_params(self) -> SearchParams:
        return SearchParams(
            date_from=self.ed_date_from.date().toString("dd.MM.yyyy"),
            query=self.ed_query.text().strip(),
            tag_id=self.ed_tag_id.value() or None,
        )

    def client_filters(self) -> ClientFilters:
        return ClientFilters(
            registry_contains=self.ed_registry.text().strip(),
            organizer_contains=self.ed_organizer.text().strip(),
            title_contains=self.ed_title_local.text().strip(),
            trend_pur=self.cb_trend.currentData() or "",
            step_id=self.cb_step.currentData() or "",
            applics_min=(self.sb_apc_min.value() or None),
            applics_max=(self.sb_apc_max.value() or None),
            price_min=(self.sb_price_min.value() or None),
            price_max=(self.sb_price_max.value() or None),
            end_from=(
                datetime.combine(self.de_end_from.date().toPython(), datetime.min.time())
                if self.cb_end_from_en.isChecked() else None
            ),
            end_to=(
                datetime.combine(self.de_end_to.date().toPython(), datetime.max.time())
                if self.cb_end_to_en.isChecked() else None
            ),
        )

    def reset_client_filters(self) -> None:
        self.ed_registry.clear()
        self.ed_organizer.clear()
        self.ed_title_local.clear()
        self.cb_trend.setCurrentIndex(0)
        self.cb_step.setCurrentIndex(0)
        self.sb_apc_min.setValue(0)
        self.sb_apc_max.setValue(0)
        self.sb_price_min.setValue(0.0)
        self.sb_price_max.setValue(0.0)
        self.cb_end_from_en.setChecked(False)
        self.cb_end_to_en.setChecked(False)

    def set_controls_enabled(self, enabled: bool) -> None:
        self.btn_search.setEnabled(enabled)


# -----------------------------------------------------------------------------
# Главное окно
# -----------------------------------------------------------------------------

class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.resize(1500, 900)

        self.client = EtpClient()
        self.runner = TaskRunner(self)
        self.model = ProcedureTableModel(self)
        self.proxy = ProcedureFilterProxy(self)
        self.proxy.setSourceModel(self.model)

        self._last_total: int = 0
        self._current_start: int = 0
        self._last_user: Optional[str] = None
        self._cache_dirty: bool = False

        self._cache_save_timer = QTimer(self)
        self._cache_save_timer.setSingleShot(True)
        self._cache_save_timer.timeout.connect(self._save_cache_now)

        self._build_ui()
        self._announce_cache_on_start()
        self._update_controls()

    # ------------------------------------------------------------------ UI
    def _build_ui(self) -> None:
        # ---------- Верхняя панель ----------
        top = QFrame()
        top.setObjectName("TopBar")
        top.setFixedHeight(56)
        top_layout = QHBoxLayout(top)
        top_layout.setContentsMargins(16, 8, 16, 8)
        top_layout.setSpacing(12)

        t_title_box = QVBoxLayout()
        t_title_box.setSpacing(0)
        self.title_label = QLabel("ЭТП ГПБ — Актуальные процедуры")
        self.title_label.setObjectName("TopBarTitle")
        self.subtitle_label = QLabel("Поиск, фильтры и экспорт")
        self.subtitle_label.setObjectName("TopBarSubtitle")
        t_title_box.addWidget(self.title_label)
        t_title_box.addWidget(self.subtitle_label)
        top_layout.addLayout(t_title_box)
        top_layout.addStretch(1)

        self.user_label = QLabel("Пользователь: —")
        self.user_label.setStyleSheet("color: #4a515a;")
        top_layout.addWidget(self.user_label)

        self.session_badge = QLabel("○  Chrome не запущен")
        self.session_badge.setObjectName("SessionBadge")
        self.session_badge.setProperty("ok", "idle")
        top_layout.addWidget(self.session_badge)

        # ---------- Сайдбар и основная область ----------
        self.sidebar = Sidebar()
        self.sidebar.searchRequested.connect(self._on_search)
        self.sidebar.resetRequested.connect(self._on_reset_filters)
        self.sidebar.clientFiltersChanged.connect(self._on_filters_changed)

        main_area = QWidget()
        main_area_layout = QVBoxLayout(main_area)
        main_area_layout.setContentsMargins(12, 10, 12, 8)
        main_area_layout.setSpacing(8)

        # Верхняя полоска со счётчиком
        actions = QFrame()
        actions_layout = QHBoxLayout(actions)
        actions_layout.setContentsMargins(0, 0, 0, 0)
        actions_layout.setSpacing(8)

        self.lbl_counter = QLabel("Данных нет. Нажмите «Поиск».")
        self.lbl_counter.setStyleSheet("color: #3a4048; font-weight: 600;")
        actions_layout.addWidget(self.lbl_counter)
        actions_layout.addStretch(1)

        self.btn_export = QPushButton("Экспорт в XLSX…")
        self.btn_export.clicked.connect(self._on_export)
        actions_layout.addWidget(self.btn_export)

        main_area_layout.addWidget(actions)

        # Таблица
        self.table = QTableView()
        self.table.setModel(self.proxy)
        self.table.setSortingEnabled(True)
        self.table.sortByColumn(6, Qt.AscendingOrder)
        self.table.setSelectionBehavior(QTableView.SelectRows)
        self.table.setSelectionMode(QTableView.ExtendedSelection)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setDefaultSectionSize(26)
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Interactive)
        hh.setSectionResizeMode(3, QHeaderView.Stretch)
        self.table.setColumnWidth(0, 170)
        self.table.setColumnWidth(1, 180)
        self.table.setColumnWidth(2, 230)
        self.table.setColumnWidth(4, 80)
        self.table.setColumnWidth(5, 95)
        self.table.setColumnWidth(6, 145)
        self.table.setColumnWidth(7, 170)
        self.table.setColumnWidth(8, 200)
        self.table.setWordWrap(False)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._on_context_menu)
        self.table.doubleClicked.connect(self._on_row_double_clicked)
        main_area_layout.addWidget(self.table, 1)

        bottom_bar = QFrame()
        bottom_bar.setObjectName("BottomBar")
        bottom_layout = QHBoxLayout(bottom_bar)
        bottom_layout.setContentsMargins(10, 8, 10, 8)
        bottom_layout.setSpacing(8)

        self.btn_load_more = QPushButton("Следующий батч")
        self.btn_load_more.setToolTip("Загрузить следующую пачку (25 процедур)")
        self.btn_load_more.clicked.connect(self._on_load_more)
        self.btn_load_more.setEnabled(False)
        bottom_layout.addWidget(self.btn_load_more)

        self.btn_load_all = QPushButton("Загрузить все батчи")
        self.btn_load_all.setToolTip("Подряд скачать все оставшиеся пачки")
        self.btn_load_all.clicked.connect(self._on_load_all)
        self.btn_load_all.setEnabled(False)
        bottom_layout.addWidget(self.btn_load_all)

        self.btn_stop = QPushButton("Стоп")
        self.btn_stop.setObjectName("Danger")
        self.btn_stop.clicked.connect(self._on_stop)
        self.btn_stop.setEnabled(False)
        bottom_layout.addWidget(self.btn_stop)

        bottom_layout.addStretch(1)
        main_area_layout.addWidget(bottom_bar)

        # Splitter
        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(self.sidebar)
        splitter.addWidget(main_area)
        splitter.setChildrenCollapsible(False)
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        splitter.setSizes([420, 1080])

        # Центральный виджет
        central = QWidget()
        cl = QVBoxLayout(central)
        cl.setContentsMargins(0, 0, 0, 0)
        cl.setSpacing(0)
        cl.addWidget(top)
        cl.addWidget(splitter, 1)
        self.setCentralWidget(central)

        # Status bar
        sb = QStatusBar()
        self.setStatusBar(sb)
        self.status_msg = QLabel("Готов.")
        sb.addWidget(self.status_msg, 1)
        self.progress = QProgressBar()
        self.progress.setFixedWidth(220)
        self.progress.setRange(0, 0)
        self.progress.hide()
        sb.addPermanentWidget(self.progress)

        # Горячие клавиши
        act_search = QAction(self)
        act_search.setShortcut(QKeySequence("Ctrl+Return"))
        act_search.triggered.connect(self._on_search)
        self.addAction(act_search)

        act_focus_query = QAction(self)
        act_focus_query.setShortcut(QKeySequence("Ctrl+F"))
        act_focus_query.triggered.connect(lambda: self.sidebar.ed_query.setFocus())
        self.addAction(act_focus_query)

    # ------------------------------------------------------------------ задачи
    def _on_search(self) -> None:
        if self.runner.is_running():
            return

        if CACHE_FILE.exists():
            choice = self._ask_cache_choice()
            if choice == "cancel":
                return
            if choice == "cache":
                self._use_cache()
                return
            # choice == "refresh" → удаляем кэш и идём парсить заново
            self._delete_cache()

        self.model.clear()
        self._current_start = 0
        self._last_total = 0
        self._refresh_counter()
        self._start_task(self.sidebar.search_params(), start=0, batches=1)

    def _ask_cache_choice(self) -> str:
        """Диалог «Показать из кэша / Загрузить заново / Отмена».

        Возвращает одно из: 'cache', 'refresh', 'cancel'.
        """
        meta = self._read_cache_meta()
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Question)
        box.setWindowTitle("Найден сохранённый результат")
        if meta:
            saved_at = (meta.get("saved_at") or "")[:16].replace("T", " ")
            count = meta.get("count") or 0
            box.setText(
                f"В кэше уже есть <b>{count}</b> процедур, сохранённых "
                f"<b>{saved_at}</b>."
            )
        else:
            box.setText("В кэше уже есть сохранённый результат поиска.")
        box.setInformativeText(
            "Что делать?\n\n"
            "• Показать из кэша — мгновенно вывести сохранённые данные.\n"
            "• Загрузить заново — очистить кэш и спарсить с сайта."
        )
        btn_cache = box.addButton("Показать из кэша", QMessageBox.AcceptRole)
        btn_refresh = box.addButton("Загрузить заново", QMessageBox.DestructiveRole)
        btn_cancel = box.addButton("Отмена", QMessageBox.RejectRole)
        box.setDefaultButton(btn_cache)
        box.exec()
        clicked = box.clickedButton()
        if clicked is btn_cache:
            return "cache"
        if clicked is btn_refresh:
            return "refresh"
        return "cancel"

    def _read_cache_meta(self) -> Optional[dict[str, Any]]:
        try:
            data = json.loads(CACHE_FILE.read_text(encoding="utf-8"))
            return {
                "saved_at": data.get("saved_at"),
                "count": len(data.get("procedures") or []),
                "total": data.get("total"),
            }
        except Exception:
            return None

    def _use_cache(self) -> None:
        """Загрузить результат из кэша в таблицу, без обращения к сайту."""
        if not CACHE_FILE.exists():
            return
        try:
            data = json.loads(CACHE_FILE.read_text(encoding="utf-8"))
        except Exception as e:
            QMessageBox.warning(
                self, "Кэш повреждён",
                f"Не удалось прочитать кэш:\n{e}\n\nБудет выполнен новый поиск.",
            )
            self._delete_cache()
            self.model.clear()
            self._current_start = 0
            self._last_total = 0
            self._start_task(self.sidebar.search_params(), start=0, batches=1)
            return
        procs = data.get("procedures") or []
        self.model.set_rows(procs)
        self._last_total = int(data.get("total") or len(procs))
        self._current_start = len(procs)
        saved_at = (data.get("saved_at") or "")[:16].replace("T", " ")
        self.status_msg.setText(
            f"Показано из кэша: {len(procs)} процедур (сохранено {saved_at})."
        )
        self._refresh_counter()
        self._update_controls()

    def _delete_cache(self) -> None:
        try:
            if CACHE_FILE.exists():
                CACHE_FILE.unlink()
        except Exception:
            traceback.print_exc()
        self._cache_dirty = False

    def _on_load_more(self) -> None:
        if self.runner.is_running():
            return
        self._start_task(self.sidebar.search_params(), start=self._current_start, batches=1)

    def _on_load_all(self) -> None:
        if self.runner.is_running():
            return
        self._start_task(self.sidebar.search_params(), start=self._current_start, batches=10_000)

    def _start_task(self, params: SearchParams, start: int, batches: int) -> None:
        self.progress.show()
        self.btn_stop.setEnabled(batches > 1)
        self.sidebar.set_controls_enabled(False)
        self.btn_load_more.setEnabled(False)
        self.btn_load_all.setEnabled(False)
        self._set_badge("idle", "● Работаю…")

        fn = _make_search_task(self.client, params, start, batches)
        try:
            self.runner.start(
                fn,
                on_progress=self._on_progress,
                on_session=self._on_session_status,
                on_batch=self._on_batch_loaded,
                on_error=self._on_error,
                on_done=self._on_task_done,
            )
        except Exception as e:
            self._on_error(f"Не удалось запустить задачу: {e}")
            self._on_task_done()

    # --------------- слоты от Worker
    @Slot(str)
    def _on_progress(self, text: str) -> None:
        self.status_msg.setText(text)

    @Slot(bool, str)
    def _on_session_status(self, ok: bool, message: str) -> None:
        if ok:
            self._set_badge("true", "● Сессия активна")
            # Подхватим логин пользователя
            try:
                login = self.client.current_user_login()
            except Exception:
                login = None
            if login:
                self._last_user = login
                self.user_label.setText(f"Пользователь: {login}")
        else:
            self._set_badge("false", "○ Нужен вход")
            # Диалог с подсказкой + кнопкой «Повторить»
            box = QMessageBox(self)
            box.setIcon(QMessageBox.Information)
            box.setWindowTitle("Требуется авторизация")
            box.setText("Сессия не активна.")
            box.setInformativeText(message)
            btn_retry = box.addButton("Я вошёл — повторить", QMessageBox.AcceptRole)
            box.addButton("Отмена", QMessageBox.RejectRole)
            box.exec()
            if box.clickedButton() is btn_retry:
                QTimer.singleShot(200, self._on_search)

    @Slot(list, int, int)
    def _on_batch_loaded(self, procs: list, start: int, total: int) -> None:
        self._last_total = total or self._last_total
        if start == 0 and self.model.rowCount() == 0:
            self.model.set_rows(procs)
        else:
            self.model.append_rows(procs)
        self._current_start = start + len(procs)
        self._refresh_counter()
        self._schedule_cache_save()

    @Slot(str)
    def _on_error(self, msg: str) -> None:
        # Не даём приложению упасть — только сообщаем пользователю.
        self._set_badge("false", "⚠ Ошибка")
        self.status_msg.setText(msg.splitlines()[0] if msg else "Ошибка")
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Warning)
        box.setWindowTitle("Ошибка")
        box.setText("При выполнении операции возникла ошибка:")
        box.setDetailedText(msg)
        box.exec()

    @Slot()
    def _on_task_done(self) -> None:
        self.progress.hide()
        self.btn_stop.setEnabled(False)
        self.sidebar.set_controls_enabled(True)
        self._update_controls()
        if self.model.rowCount() > 0:
            self.status_msg.setText(
                f"Загружено {self.model.rowCount()} / {self._last_total or self.model.rowCount()} процедур."
            )

    def _on_stop(self) -> None:
        self.runner.request_stop()
        self.status_msg.setText("Останавливаю…")

    # --------------- клиентские фильтры
    def _on_filters_changed(self) -> None:
        self.proxy.set_filters(self.sidebar.client_filters())
        self._refresh_counter()

    def _on_reset_filters(self) -> None:
        self.sidebar.reset_client_filters()
        self.proxy.set_filters(ClientFilters())
        self._refresh_counter()

    # --------------- таблица
    def _proc_from_index(self, idx: QModelIndex) -> Optional[dict[str, Any]]:
        if not idx.isValid():
            return None
        src = self.proxy.mapToSource(idx)
        return self.model.row_at(src.row())

    def _on_row_double_clicked(self, idx: QModelIndex) -> None:
        self._open_in_browser(self._proc_from_index(idx))

    def _on_context_menu(self, pos) -> None:
        idx = self.table.indexAt(pos)
        if not idx.isValid():
            return
        proc = self._proc_from_index(idx)
        menu = QMenu(self)
        menu.addAction("Открыть в Chrome", lambda: self._open_in_browser(proc))
        menu.addSeparator()
        menu.addAction(
            "Копировать реестровый №",
            lambda: QApplication.clipboard().setText(
                str((proc or {}).get("registry_number") or (proc or {}).get("procedure_number") or "")
            ),
        )
        menu.addAction(
            "Копировать наименование",
            lambda: QApplication.clipboard().setText(str((proc or {}).get("title") or "")),
        )
        menu.addAction(
            "Копировать ИНН организатора",
            lambda: QApplication.clipboard().setText(str((proc or {}).get("org_inn") or "")),
        )
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def _open_in_browser(self, proc: Optional[dict[str, Any]]) -> None:
        if not proc:
            return
        pid = proc.get("id")
        if not pid:
            return
        webbrowser.open(VIEW_URL.format(pid=pid))

    # --------------- экспорт
    def _on_export(self) -> None:
        if self.model.rowCount() == 0:
            QMessageBox.information(self, "Нет данных", "Сначала выполните поиск.")
            return
        default_name = f"procedures_{datetime.now():%Y%m%d_%H%M}.xlsx"
        path_str, _ = QFileDialog.getSaveFileName(
            self, "Сохранить в XLSX",
            str(Path.cwd() / default_name),
            "Excel (*.xlsx)",
        )
        if not path_str:
            return
        visible: list[dict[str, Any]] = []
        for i in range(self.proxy.rowCount()):
            src = self.proxy.mapToSource(self.proxy.index(i, 0))
            r = self.model.row_at(src.row())
            if r is not None:
                visible.append(r)
        try:
            self._write_xlsx(Path(path_str), visible)
            QMessageBox.information(
                self, "Готово",
                f"Сохранено {len(visible)} строк в\n{path_str}",
            )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка экспорта", str(e))

    def _write_xlsx(self, path: Path, procs: list[dict[str, Any]]) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Процедуры"
        titles = [c[1] for c in COLUMNS] + ["id", "Дата публикации"]
        ws.append(titles)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for p in procs:
            ws.append(
                [
                    p.get("registry_number") or p.get("procedure_number") or "",
                    trend_pur_label(p.get("trend_pur")),
                    p.get("short_name") or p.get("full_name") or "",
                    p.get("title") or "",
                    ", ".join(str(t) for t in (p.get("tags") or [])),
                    p.get("applics_count") or 0,
                    _fmt_date(_parse_dt(p.get("date_end_registration"))),
                    _parse_price(p.get("total_price")) or "",
                    step_id_label(p.get("step_id")),
                    p.get("id"),
                    _fmt_date(_parse_dt(p.get("date_published"))),
                ]
            )
        widths = [18, 28, 28, 80, 12, 12, 20, 18, 28, 10, 20]
        for i, w in enumerate(widths, start=1):
            col_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[col_letter].width = w
        ws.freeze_panes = "A2"
        wb.save(path)

    # --------------- счётчики / бейдж
    def _refresh_counter(self) -> None:
        loaded = self.model.rowCount()
        visible = self.proxy.rowCount()
        total = self._last_total
        if loaded == 0 and total == 0:
            self.lbl_counter.setText("Данных нет. Нажмите «Поиск».")
        elif total and loaded < total:
            self.lbl_counter.setText(
                f"Показано {visible} (загружено {loaded}) из {total} по фильтру поиска."
            )
        else:
            self.lbl_counter.setText(f"Показано {visible} из {loaded} процедур.")
        self._update_controls()

    def _set_badge(self, state: str, text: str) -> None:
        self.session_badge.setProperty("ok", state)
        self.session_badge.setText(text)
        self.session_badge.style().unpolish(self.session_badge)
        self.session_badge.style().polish(self.session_badge)

    def _update_controls(self) -> None:
        running = self.runner.is_running()
        loaded = self.model.rowCount()
        total = self._last_total
        has_more = total > 0 and loaded < total
        self.btn_load_more.setEnabled(not running and has_more)
        self.btn_load_all.setEnabled(not running and has_more)
        self.btn_export.setEnabled(loaded > 0)
        self.sidebar.set_controls_enabled(not running)

    # --------------- кэш
    def _schedule_cache_save(self) -> None:
        self._cache_dirty = True
        self._cache_save_timer.start(1000)

    def _save_cache_now(self) -> None:
        if not self._cache_dirty:
            return
        self._cache_dirty = False
        try:
            CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
            payload = {
                "saved_at": datetime.now().isoformat(),
                "date_from": self.sidebar.ed_date_from.date().toString("dd.MM.yyyy"),
                "query": self.sidebar.ed_query.text().strip(),
                "total": self._last_total,
                "procedures": self.model.rows(),
            }
            CACHE_FILE.write_text(
                json.dumps(payload, ensure_ascii=False), encoding="utf-8"
            )
        except Exception:
            traceback.print_exc()

    def _announce_cache_on_start(self) -> None:
        """При старте только сообщаем, что есть кэш — не загружаем его автоматически.

        Сам выбор (использовать / перезагрузить) делается при клике «Поиск».
        """
        meta = self._read_cache_meta()
        if meta and meta.get("count"):
            saved_at = (meta.get("saved_at") or "")[:16].replace("T", " ")
            self.lbl_counter.setText(
                f"Есть сохранённый результат: {meta['count']} процедур от {saved_at}. "
                "Нажмите «Поиск», чтобы выбрать действие."
            )
            self.status_msg.setText(
                "Готов. Найден кэш — выбор предложат при нажатии «Поиск»."
            )
        else:
            self.status_msg.setText("Готов. Нажмите «Поиск».")

    # --------------- закрытие
    def closeEvent(self, event) -> None:  # noqa: N802
        try:
            self._save_cache_now()
        except Exception:
            pass
        try:
            self.runner.shutdown(wait_ms=2000)
        except Exception:
            pass
        try:
            self.client.close()
        except Exception:
            pass
        event.accept()


# -----------------------------------------------------------------------------
# main
# -----------------------------------------------------------------------------

def _install_global_excepthook() -> None:
    """Глобальный перехват исключений: UI не должен закрываться из-за них."""

    def hook(exc_type, exc, tb) -> None:
        msg = "".join(traceback.format_exception(exc_type, exc, tb))
        print(msg, file=sys.stderr)
        try:
            QMessageBox.critical(None, "Необработанная ошибка", msg)
        except Exception:
            pass

    sys.excepthook = hook


def main() -> int:
    _install_global_excepthook()
    app = QApplication(sys.argv)
    app.setApplicationName(APP_TITLE)
    app.setQuitOnLastWindowClosed(True)

    font = QFont()
    font.setFamily("Segoe UI")
    font.setPointSize(9)
    app.setFont(font)
    palette = QPalette()
    palette.setColor(QPalette.Window, QColor("#f7f8fa"))
    palette.setColor(QPalette.WindowText, QColor("#1a1d22"))
    palette.setColor(QPalette.Base, QColor("#ffffff"))
    palette.setColor(QPalette.AlternateBase, QColor("#f8f9fb"))
    palette.setColor(QPalette.Text, QColor("#1a1d22"))
    palette.setColor(QPalette.Button, QColor("#ffffff"))
    palette.setColor(QPalette.ButtonText, QColor("#1a1d22"))
    palette.setColor(QPalette.Highlight, QColor("#cfe0ff"))
    palette.setColor(QPalette.HighlightedText, QColor("#1a1d22"))
    palette.setColor(QPalette.ToolTipBase, QColor("#2c2f36"))
    palette.setColor(QPalette.ToolTipText, QColor("#ffffff"))
    app.setPalette(palette)
    app.setStyleSheet(APP_STYLE)

    win = MainWindow()
    win.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
