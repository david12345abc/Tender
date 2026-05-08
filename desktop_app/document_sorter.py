from __future__ import annotations

import csv
import re
import shutil
from pathlib import Path
from typing import Any

COMMERCIAL_DIR_NAME = "Коммерческие"
TECHNICAL_DIR_NAME = "Технические"

PRICE_CURRENCY_RE = re.compile(
    r"(?:цена|стоимость|сумма|итого|нмц)[^\n\r]{0,80}"
    r"(?:руб\.?|rur|rub|₽|usd|eur|cny|kzt|тенге|юан[ьяей]?)",
    re.IGNORECASE,
)


def _safe_target_path(target_dir: Path, filename: str) -> Path:
    target = target_dir / filename
    if not target.exists():
        return target
    stem = target.stem
    suffix = target.suffix
    n = 2
    while True:
        candidate = target_dir / f"{stem}_{n}{suffix}"
        if not candidate.exists():
            return candidate
        n += 1


def _flatten_values(values: list[Any]) -> str:
    return " ".join(str(v).strip() for v in values if v not in (None, ""))


def _read_xlsx_headers(path: Path, max_rows: int = 25, max_cols: int = 30) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    chunks: list[str] = []
    try:
        for sheet in workbook.worksheets:
            for row_index, row in enumerate(
                sheet.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True),
                start=1,
            ):
                if row_index > max_rows:
                    break
                chunks.append(_flatten_values(list(row)))
    finally:
        workbook.close()
    return "\n".join(chunks)


def _read_xls_headers_via_excel(path: Path, max_rows: int = 25, max_cols: int = 30) -> str:
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    chunks: list[str] = []
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(path), ReadOnly=True)
        for sheet in workbook.Worksheets:
            used = sheet.UsedRange
            row_count = min(int(used.Rows.Count), max_rows)
            col_count = min(int(used.Columns.Count), max_cols)
            for row in range(1, row_count + 1):
                values = [sheet.Cells(row, col).Text for col in range(1, col_count + 1)]
                chunks.append(_flatten_values(values))
    finally:
        if workbook is not None:
            workbook.Close(False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()
    return "\n".join(chunks)


def _read_docx_tables(path: Path, max_rows: int = 25) -> str:
    from docx import Document

    doc = Document(path)
    chunks: list[str] = []
    for table in doc.tables:
        for row_index, row in enumerate(table.rows, start=1):
            if row_index > max_rows:
                break
            chunks.append(_flatten_values([cell.text for cell in row.cells]))
    return "\n".join(chunks)


def _read_csv_headers(path: Path, max_rows: int = 25) -> str:
    raw = path.read_bytes()
    text = ""
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = raw.decode(encoding, errors="replace")
            break
        except Exception:
            continue
    if not text:
        text = raw.decode("utf-8", errors="replace")
    sample = text[:4096]
    dialect = csv.Sniffer().sniff(sample, delimiters=";,	,")
    rows: list[str] = []
    for index, row in enumerate(csv.reader(text.splitlines(), dialect), start=1):
        if index > max_rows:
            break
        rows.append(_flatten_values(row))
    return "\n".join(rows)


def _document_probe_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx_headers(path)
    if suffix == ".xls":
        return _read_xls_headers_via_excel(path)
    if suffix == ".docx":
        return _read_docx_tables(path)
    if suffix in {".csv", ".tsv"}:
        return _read_csv_headers(path)
    return ""


def is_commercial_document(path: Path) -> bool:
    probe_text = _document_probe_text(path)
    return bool(PRICE_CURRENCY_RE.search(probe_text))


def sort_filled_documents(folder: Path) -> dict[str, Any]:
    commercial_dir = folder / COMMERCIAL_DIR_NAME
    technical_dir = folder / TECHNICAL_DIR_NAME
    commercial_dir.mkdir(parents=True, exist_ok=True)
    technical_dir.mkdir(parents=True, exist_ok=True)

    result: dict[str, Any] = {
        "commercial": [],
        "technical": [],
        "errors": [],
        "folder": str(folder),
    }

    for path in sorted(folder.iterdir(), key=lambda item: item.name.casefold()):
        if path.name in {COMMERCIAL_DIR_NAME, TECHNICAL_DIR_NAME}:
            continue
        if path.is_dir() or path.name.startswith("~$"):
            continue
        try:
            target_dir = commercial_dir if is_commercial_document(path) else technical_dir
            target = _safe_target_path(target_dir, path.name)
            shutil.move(str(path), str(target))
            key = "commercial" if target_dir == commercial_dir else "technical"
            result[key].append(str(target))
        except Exception as e:
            result["errors"].append(f"{path.name}: {e}")

    return result
