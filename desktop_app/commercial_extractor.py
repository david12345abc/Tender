from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from desktop_app.constants import LM_STUDIO_BASE_URL, LM_STUDIO_MODEL
from desktop_app.document_text import build_documents_text
from desktop_app.lm_table_analysis import call_lm_studio_chat


SUPPORTED_SUFFIXES = {
    ".xlsx",
    ".xlsm",
    ".xls",
    ".docx",
    ".doc",
    ".pdf",
    ".txt",
    ".csv",
    ".rtf",
    ".odt",
    ".ods",
}

MONTHS_RU = {
    "января": 1,
    "февраля": 2,
    "марта": 3,
    "апреля": 4,
    "мая": 5,
    "июня": 6,
    "июля": 7,
    "августа": 8,
    "сентября": 9,
    "октября": 10,
    "ноября": 11,
    "декабря": 12,
}


@dataclass
class CommercialTerms:
    price_with_vat: Decimal | None = None
    price_without_vat: Decimal | None = None
    currency: str = "RUB"
    validity_date: str = ""
    confidence: float = 0.0
    source: str = "rules"
    notes: list[str] = field(default_factory=list)
    files: list[str] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        return {
            "price_with_vat": _format_decimal(self.price_with_vat),
            "price_without_vat": _format_decimal(self.price_without_vat),
            "currency": self.currency,
            "validity_date": self.validity_date,
            "confidence": round(float(self.confidence), 2),
            "source": self.source,
            "notes": self.notes,
            "files": self.files,
        }


def extract_commercial_terms(
    commercial_dir: Path,
    progress=None,
    use_llm: bool = True,
) -> CommercialTerms:
    """Извлекает итоговую цену и срок действия предложения из папки Коммерческие."""
    files = _collect_files(commercial_dir)
    terms = CommercialTerms(files=[str(path) for path in files])
    if not files:
        terms.notes.append(f"В папке коммерческих документов нет поддерживаемых файлов: {commercial_dir}")
        return terms
    terms.notes.append(f"Папка анализа: {commercial_dir}")
    terms.notes.append(f"Найдено поддерживаемых файлов: {len(files)}")
    for path in files:
        terms.notes.append(f"Файл: {path.name}")

    if progress:
        progress("Анализирую коммерческие документы правилами...")

    candidates: list[dict[str, Any]] = []
    text_parts: list[str] = []
    for path in files:
        try:
            if path.suffix.lower() in {".xlsx", ".xlsm"}:
                xlsx_candidates = _price_candidates_from_xlsx(path)
                candidates.extend(xlsx_candidates)
                terms.notes.append(f"{path.name}: кандидатов цены из Excel: {len(xlsx_candidates)}")
            elif path.suffix.lower() == ".xls":
                xls_candidates = _price_candidates_from_xls(path)
                candidates.extend(xls_candidates)
                terms.notes.append(f"{path.name}: кандидатов цены из XLS: {len(xls_candidates)}")
        except Exception as e:
            terms.notes.append(f"{path.name}: не удалось разобрать таблицу Excel ({e})")

    try:
        text = build_documents_text(files, progress=None)
    except Exception as e:
        text = ""
        terms.notes.append(f"Не удалось извлечь общий текст коммерческих документов: {e}")
    if text:
        text_parts.append(text)
        text_candidates = _price_candidates_from_text(text)
        candidates.extend(text_candidates)
        terms.notes.append(f"Кандидатов цены из общего текста: {len(text_candidates)}")
        terms.validity_date = _extract_validity_date(text)
        if terms.validity_date:
            terms.notes.append(f"Дата действия найдена правилами: {terms.validity_date}")
        else:
            terms.notes.append("Дата действия правилами не найдена.")

    best = _choose_best_price_candidate(candidates)
    if candidates:
        terms.notes.append("Кандидаты цены:")
        for index, candidate in enumerate(sorted(candidates, key=lambda item: float(item.get("score") or 0), reverse=True)[:8], start=1):
            terms.notes.append(f"{index}. {_candidate_log_line(candidate)}")
    else:
        terms.notes.append("Кандидаты цены правилами не найдены.")

    if best:
        value = best["value"]
        vat_kind = best.get("vat_kind") or ""
        if vat_kind == "without_vat":
            terms.price_without_vat = value
            terms.price_with_vat = _add_vat(value)
        else:
            terms.price_with_vat = value
            terms.price_without_vat = _remove_vat(value)
        terms.currency = best.get("currency") or "RUB"
        terms.confidence = float(best.get("score") or 0)
        terms.notes.append(f"Цена найдена правилами: {best.get('reason', 'кандидат с лучшим весом')}")
        terms.notes.append(f"Выбрана цена: {_format_decimal(terms.price_with_vat)} с НДС, {_format_decimal(terms.price_without_vat)} без НДС")

    needs_llm = (
        use_llm
        and text_parts
        and (terms.price_with_vat is None or terms.confidence < 0.82 or not terms.validity_date)
    )
    terms.notes.append(f"Нужно уточнение через LM Studio: {'да' if needs_llm else 'нет'}")
    if needs_llm:
        if progress:
            progress("Правила дали не все значения, уточняю через LM Studio...")
        try:
            llm_terms = _extract_with_lm_studio("\n\n".join(text_parts))
            terms = _merge_llm_terms(terms, llm_terms)
            terms.notes.append(
                "После LM Studio: "
                f"с НДС={_format_decimal(terms.price_with_vat) or '-'}, "
                f"без НДС={_format_decimal(terms.price_without_vat) or '-'}, "
                f"дата={terms.validity_date or '-'}, "
                f"уверенность={terms.confidence:.2f}"
            )
        except Exception as e:
            terms.notes.append(f"LM Studio не смог уточнить коммерческие условия: {e}")

    if terms.price_with_vat is not None and terms.price_without_vat is None:
        terms.price_without_vat = _remove_vat(terms.price_with_vat)
    if terms.price_without_vat is not None and terms.price_with_vat is None:
        terms.price_with_vat = _add_vat(terms.price_without_vat)

    return terms


def _candidate_log_line(candidate: dict[str, Any]) -> str:
    value = candidate.get("value")
    amount = _format_decimal(value if isinstance(value, Decimal) else None) or str(value or "")
    return (
        f"{amount} {candidate.get('currency') or 'RUB'}, "
        f"score={float(candidate.get('score') or 0):.2f}, "
        f"НДС={candidate.get('vat_kind') or 'не определено'}, "
        f"источник={candidate.get('reason') or '-'}"
    )


def _collect_files(root: Path) -> list[Path]:
    if not root.exists():
        return []
    return sorted(
        (
            path
            for path in root.rglob("*")
            if path.is_file()
            and not path.name.startswith("~$")
            and path.suffix.lower() in SUPPORTED_SUFFIXES
        ),
        key=lambda path: str(path).casefold(),
    )


def _price_candidates_from_xlsx(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    candidates: list[dict[str, Any]] = []
    try:
        for ws in wb.worksheets[:8]:
            rows = list(ws.iter_rows(values_only=True))
            candidates.extend(_price_candidates_from_table_rows(path.name, ws.title, rows))
            for row_index, row in enumerate(rows, start=1):
                cells = ["" if value is None else str(value).strip() for value in row]
                row_text = " ".join(cells).lower()
                numbers = [_parse_amount(cell) for cell in cells]
                values = [value for value in numbers if value is not None and value > Decimal("0")]
                if values and _looks_like_total_row(row_text):
                    candidates.append(
                        {
                            "value": max(values),
                            "score": 0.95,
                            "currency": _detect_currency(row_text),
                            "vat_kind": _detect_vat_kind(row_text),
                            "reason": f"{path.name}, лист {ws.title}, строка с итогом {row_index}",
                        }
                    )
                    continue

                if not values:
                    continue
                header_text = _nearby_header_text(rows, row_index - 1)
                combined = f"{header_text} {row_text}".lower()
                if _looks_like_sum_context(combined):
                    score = 0.72
                    if _has_currency(combined):
                        score += 0.08
                    candidates.append(
                        {
                            "value": max(values),
                            "score": min(score, 0.8),
                            "currency": _detect_currency(combined),
                            "vat_kind": _detect_vat_kind(combined),
                            "reason": f"{path.name}, лист {ws.title}, строка {row_index}",
                        }
                    )
    finally:
        wb.close()
    return candidates


def _price_candidates_from_xls(path: Path) -> list[dict[str, Any]]:
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(path), ReadOnly=True)
        try:
            candidates: list[dict[str, Any]] = []
            for sheet in list(wb.Worksheets)[:8]:
                used = sheet.UsedRange
                values = used.Value
                rows = _normalize_excel_rows(values)
                candidates.extend(_price_candidates_from_table_rows(path.name, str(sheet.Name), rows))
            return candidates
        finally:
            wb.Close(False)
    finally:
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def _normalize_excel_rows(values: Any) -> list[tuple[Any, ...]]:
    if values is None:
        return []
    if not isinstance(values, tuple):
        return [(values,)]
    if values and not isinstance(values[0], tuple):
        return [values]
    return list(values)


def _price_candidates_from_table_rows(
    file_name: str,
    sheet_name: str,
    rows: list[tuple[Any, ...]],
) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    if not rows:
        return candidates
    max_cols = max((len(row) for row in rows), default=0)
    for col_index in range(max_cols):
        header_parts: list[str] = []
        for row in rows[:12]:
            if col_index < len(row) and row[col_index] is not None:
                text = str(row[col_index]).strip()
                if text:
                    header_parts.append(text)
        header = " ".join(header_parts)
        score = _total_price_header_score(header)
        if score <= 0:
            continue
        amounts: list[Decimal] = []
        for row in rows[1:]:
            if col_index >= len(row):
                continue
            amount = _parse_cell_amount(row[col_index])
            if amount is not None and amount > Decimal("0"):
                amounts.append(amount)
        if not amounts:
            continue
        total = sum(amounts, Decimal("0")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        candidates.append(
            {
                "value": total,
                "score": score,
                "currency": _detect_currency(header),
                "vat_kind": _detect_vat_kind(header) or "with_vat",
                "reason": f"{file_name}, лист {sheet_name}, сумма колонки: {header[:140]}",
            }
        )
    return candidates


def _total_price_header_score(header: str) -> float:
    text = header.lower()
    if not re.search(r"(стоимость|сумма|итог)", text, re.I):
        return 0.0
    if re.search(r"(цена\s+за\s+ед|за\s+единиц|единицу|стоимость\s+шмр\s+за\s+единицу|стоимость\s+пнр\s+за\s+единицу)", text, re.I):
        return 0.0
    if "итоговая стоимость" in text and re.search(r"с\s+ндс", text):
        return 0.99
    if re.search(r"стоимость\s+с\s+ндс", text):
        return 0.93
    if re.search(r"стоимость\s+без\s+ндс", text):
        return 0.86
    if "итоговая" in text:
        return 0.9
    return 0.72


def _price_candidates_from_text(text: str) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for line in text.splitlines():
        normalized = re.sub(r"\s+", " ", line).strip()
        if not normalized:
            continue
        low = normalized.lower()
        if not (_looks_like_total_row(low) or _looks_like_sum_context(low)):
            continue
        values = _amounts_from_text(normalized)
        if not values:
            continue
        score = 0.88 if _looks_like_total_row(low) else 0.74
        if _has_currency(low):
            score += 0.05
        candidates.append(
            {
                "value": max(values),
                "score": min(score, 0.95),
                "currency": _detect_currency(low),
                "vat_kind": _detect_vat_kind(low),
                "reason": f"текстовая строка: {normalized[:160]}",
            }
        )
    return candidates


def _choose_best_price_candidate(candidates: list[dict[str, Any]]) -> dict[str, Any] | None:
    sane = [
        item
        for item in candidates
        if isinstance(item.get("value"), Decimal)
        and Decimal("1") <= item["value"] <= Decimal("999999999999")
    ]
    if not sane:
        return None
    return sorted(sane, key=lambda item: (float(item.get("score") or 0), item["value"]), reverse=True)[0]


def _looks_like_total_row(text: str) -> bool:
    return bool(
        re.search(
            r"\b(итого|всего|общая\s+стоимость|общая\s+цена|итоговая\s+стоимость|итоговая\s+цена|стоимость\s+предложения|цена\s+предложения)\b",
            text,
            re.I,
        )
    )


def _looks_like_sum_context(text: str) -> bool:
    if re.search(r"(цена\s+за\s+ед|единиц|ед\.|шт\.|кол-?во|количество)", text, re.I):
        return False
    return bool(re.search(r"(сумма|стоимость|цена).{0,40}(руб|₽|ндс|валют|итог)|итог.{0,40}(сумма|стоимость|цена)", text, re.I))


def _nearby_header_text(rows: list[tuple[Any, ...]], zero_based_row: int) -> str:
    start = max(0, zero_based_row - 4)
    parts: list[str] = []
    for row in rows[start:zero_based_row]:
        values = [str(value).strip() for value in row if value is not None and str(value).strip()]
        if values:
            parts.append(" ".join(values))
    return " ".join(parts)


def _extract_validity_date(text: str) -> str:
    compact = re.sub(r"\s+", " ", text)
    patterns = [
        r"(?:действ(?:ует|ительно|ительна)|срок\s+действия|оферт[аы]\s+действ).*?(?:до|по)\s+(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})",
        r"(?:действ(?:ует|ительно|ительна)|срок\s+действия|оферт[аы]\s+действ).*?(?:до|по)\s+(\d{1,2}\s+[а-яё]+\s+\d{4})",
        r"(?:до|по)\s+(\d{1,2}[./-]\d{1,2}[./-]\d{2,4}).{0,80}(?:действ(?:ует|ительно|ительна)|оферт|предложени)",
    ]
    for pattern in patterns:
        match = re.search(pattern, compact, re.I)
        if match:
            parsed = _normalize_date(match.group(1))
            if parsed:
                return parsed
    return ""


def _normalize_date(value: str) -> str:
    text = value.strip().lower()
    match = re.match(r"(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})$", text)
    if match:
        day, month, year = (int(part) for part in match.groups())
        if year < 100:
            year += 2000
        return _safe_date(day, month, year)
    match = re.match(r"(\d{1,2})\s+([а-яё]+)\s+(\d{4})$", text, re.I)
    if match:
        day = int(match.group(1))
        month = MONTHS_RU.get(match.group(2).lower())
        year = int(match.group(3))
        if month:
            return _safe_date(day, month, year)
    return ""


def _safe_date(day: int, month: int, year: int) -> str:
    try:
        dt = date(year, month, day)
    except ValueError:
        return ""
    return dt.strftime("%d.%m.%Y")


def _amount_re() -> re.Pattern[str]:
    return re.compile(r"(?<!\d)(?:\d{1,3}(?:[ \u00a0]\d{3})+|\d+)(?:[,.]\d{1,2})?(?!\d)")


def _parse_amount(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        except InvalidOperation:
            return None
    text = str(value)
    match = _amount_re().search(text)
    if not match:
        return None
    raw = match.group(0).replace("\u00a0", " ").replace(" ", "").replace(",", ".")
    try:
        return Decimal(raw).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except InvalidOperation:
        return None


def _parse_cell_amount(value: Any) -> Decimal | None:
    if isinstance(value, (int, float, Decimal)):
        return _parse_amount(value)
    text = str(value or "").strip()
    if not text or re.search(r"[а-яёa-z]", text, re.I):
        return None
    return _parse_amount(text)


def _amounts_from_text(text: str) -> list[Decimal]:
    values: list[Decimal] = []
    for match in _amount_re().finditer(text):
        start, end = match.span()
        left = text[max(0, start - 12):start].lower()
        right = text[end:end + 24].lower()
        if right.lstrip().startswith("%") or re.search(r"(дн|рабоч|календар|год|месяц)", right, re.I):
            continue
        if re.search(r"(ндс\s*)$", left, re.I) and right.lstrip().startswith("%"):
            continue
        amount = _parse_amount(match.group(0))
        if amount is not None and amount > Decimal("0"):
            values.append(amount)
    return values


def _detect_currency(text: str) -> str:
    value = text.lower()
    if re.search(r"\b(usd|доллар)", value):
        return "USD"
    if re.search(r"\b(eur|евро)", value):
        return "EUR"
    if re.search(r"\b(cny|юан)", value):
        return "CNY"
    return "RUB"


def _has_currency(text: str) -> bool:
    return bool(re.search(r"(руб|₽|usd|eur|cny|доллар|евро|юан)", text, re.I))


def _detect_vat_kind(text: str) -> str:
    value = text.lower()
    if re.search(r"(без\s+ндс|ндс\s+не\s+облагается|ндс\s*0)", value):
        return "without_vat"
    if re.search(r"(с\s+ндс|включая\s+ндс|ндс\s+22|ндс\s+20|ндс\s+18)", value):
        return "with_vat"
    return ""


def _add_vat(value: Decimal) -> Decimal:
    return (value * Decimal("1.22")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _remove_vat(value: Decimal) -> Decimal:
    return (value / Decimal("1.22")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _format_decimal(value: Decimal | None) -> str:
    if value is None:
        return ""
    return f"{value:.2f}"


def _extract_with_lm_studio(documents_text: str) -> CommercialTerms:
    system = (
        "Ты помогаешь заполнить письмо о подаче заявки на российской ЭТП. "
        "Из текста коммерческих документов извлеки только явно указанную итоговую стоимость предложения "
        "и дату, до которой действует предложение/оферта. Ответь строго JSON без markdown. "
        "Ключи: price_with_vat, price_without_vat, currency, validity_date, confidence, note. "
        "Цены указывай числом без пробелов, валюта RUB/USD/EUR/CNY, дата в формате ДД.ММ.ГГГГ. "
        "Если точного значения нет, верни пустую строку. confidence от 0 до 1."
    )
    user = (
        "Коммерческие документы:\n"
        "-----\n"
        f"{documents_text[:60000]}\n"
        "-----"
    )
    raw = call_lm_studio_chat(LM_STUDIO_BASE_URL, LM_STUDIO_MODEL, system, user, timeout_sec=240)
    data = _first_json(raw)
    terms = CommercialTerms(source="lm_studio")
    terms.price_with_vat = _parse_amount(data.get("price_with_vat"))
    terms.price_without_vat = _parse_amount(data.get("price_without_vat"))
    terms.currency = str(data.get("currency") or "RUB").strip().upper() or "RUB"
    terms.validity_date = _normalize_date(str(data.get("validity_date") or "")) or str(data.get("validity_date") or "").strip()
    try:
        terms.confidence = max(0.0, min(1.0, float(str(data.get("confidence") or "0").replace(",", "."))))
    except ValueError:
        terms.confidence = 0.0
    note = str(data.get("note") or "").strip()
    if note:
        terms.notes.append(f"LM Studio: {note}")
    return terms


def _first_json(raw: str) -> dict[str, Any]:
    text = raw.strip()
    match = re.match(r"^```(?:json)?\s*([\s\S]*?)\s*```$", text, re.I)
    if match:
        text = match.group(1).strip()
    decoder = json.JSONDecoder()
    for index, char in enumerate(text):
        if char not in "{[":
            continue
        try:
            value, _ = decoder.raw_decode(text[index:])
        except json.JSONDecodeError:
            continue
        if isinstance(value, list) and value and isinstance(value[0], dict):
            return value[0]
        if isinstance(value, dict):
            return value
    raise ValueError("В ответе LM Studio не найден JSON-объект.")


def _merge_llm_terms(rule_terms: CommercialTerms, llm_terms: CommercialTerms) -> CommercialTerms:
    if llm_terms.price_with_vat is not None and (
        rule_terms.price_with_vat is None or llm_terms.confidence >= rule_terms.confidence
    ):
        rule_terms.price_with_vat = llm_terms.price_with_vat
        rule_terms.source = "lm_studio"
    if llm_terms.price_without_vat is not None and (
        rule_terms.price_without_vat is None or llm_terms.confidence >= rule_terms.confidence
    ):
        rule_terms.price_without_vat = llm_terms.price_without_vat
        rule_terms.source = "lm_studio"
    if llm_terms.validity_date and not rule_terms.validity_date:
        rule_terms.validity_date = llm_terms.validity_date
        rule_terms.source = "lm_studio"
    if llm_terms.currency:
        rule_terms.currency = llm_terms.currency
    if llm_terms.confidence > rule_terms.confidence:
        rule_terms.confidence = llm_terms.confidence
    rule_terms.notes.extend(llm_terms.notes)
    return rule_terms
